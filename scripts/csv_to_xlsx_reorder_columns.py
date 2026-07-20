"""Convert CSV files to XLSX and optionally reverse a column range.

This utility can:

1. Convert one CSV file to XLSX.
2. Convert all CSV files in a folder to XLSX.
3. Optionally reverse/reorder a column range in the resulting XLSX file(s).

Examples
--------
Convert one CSV file:

    python scripts/csv_to_xlsx.py --input data/sample.csv

Convert all CSV files in a folder:

    python scripts/csv_to_xlsx.py --input data/csv_files --output data/xlsx_files

Convert and reverse Excel columns AB to AT:

    python scripts/csv_to_xlsx.py --input data/sample.csv --reverse-columns AB:AT

Reverse columns AB to AT in an existing XLSX file:

    python scripts/csv_to_xlsx.py --input data/sample.xlsx --reverse-columns AB:AT
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


def parse_column_range(value: str | None) -> tuple[int, int] | None:
    """Parse a range such as 'AB:AT' or '28:46' into 1-based column indexes."""
    if value is None or str(value).strip() == "":
        return None

    parts = str(value).strip().split(":")
    if len(parts) != 2:
        raise ValueError("Column range must use the form START:END, for example AB:AT or 28:46.")

    def parse_part(part: str) -> int:
        part = part.strip()
        if part.isdigit():
            return int(part)
        return column_index_from_string(part.upper())

    start_col = parse_part(parts[0])
    end_col = parse_part(parts[1])
    if start_col <= 0 or end_col <= 0 or start_col > end_col:
        raise ValueError(f"Invalid column range: {value}")
    return start_col, end_col


def reverse_columns_in_xlsx(xlsx_path: Path, column_range: tuple[int, int]) -> None:
    """Reverse the order of columns in an XLSX workbook over a 1-based range."""
    start_col, end_col = column_range
    workbook = load_workbook(xlsx_path)
    worksheet = workbook.active

    columns = []
    for col in range(start_col, end_col + 1):
        columns.append([
            worksheet.cell(row=row, column=col).value
            for row in range(1, worksheet.max_row + 1)
        ])

    for new_col, values in zip(range(start_col, end_col + 1), reversed(columns)):
        for row, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=new_col).value = value

    workbook.save(xlsx_path)


def convert_one_file(
    input_file: Path,
    output_file: Path | None = None,
    reverse_columns: tuple[int, int] | None = None,
) -> Path:
    """Convert one CSV file or process one existing XLSX file."""
    input_file = Path(input_file)

    if input_file.suffix.lower() == ".csv":
        if output_file is None:
            output_file = input_file.with_suffix(".xlsx")
        output_file = Path(output_file)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        df = pd.read_csv(input_file)
        df.to_excel(output_file, index=False)
    elif input_file.suffix.lower() in {".xlsx", ".xlsm"}:
        if output_file is None:
            output_file = input_file
        else:
            output_file = Path(output_file)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            output_file.write_bytes(input_file.read_bytes())
    else:
        raise ValueError(f"Unsupported input file type: {input_file.suffix}")

    if reverse_columns is not None:
        reverse_columns_in_xlsx(output_file, reverse_columns)

    return output_file


def convert_path(
    input_path: Path,
    output_path: Path | None = None,
    reverse_columns: tuple[int, int] | None = None,
) -> list[Path]:
    """Convert/process one file or all CSV/XLSX files in a folder."""
    input_path = Path(input_path)

    if input_path.is_file():
        return [convert_one_file(input_path, output_path, reverse_columns)]

    if input_path.is_dir():
        output_dir = input_path if output_path is None else Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        saved = []

        for file_path in sorted(input_path.iterdir()):
            if file_path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
                continue
            target = output_dir / f"{file_path.stem}.xlsx"
            saved.append(convert_one_file(file_path, target, reverse_columns))

        return saved

    raise FileNotFoundError(f"Input path not found: {input_path}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert CSV file(s) to XLSX and optionally reverse a column range."
    )
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        type=Path,
        help="Input CSV/XLSX file or folder containing CSV/XLSX files.",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        type=Path,
        help="Output XLSX file or output folder. Default: same location as input.",
    )
    parser.add_argument(
        "--reverse-columns",
        default=None,
        help="Optional column range to reverse, for example AB:AT or 28:46.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    reverse_range = parse_column_range(args.reverse_columns)
    saved = convert_path(args.input, args.output, reverse_range)

    for path in saved:
        print(f"Saved: {path}")

    if not saved:
        print("No CSV/XLSX files found.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
